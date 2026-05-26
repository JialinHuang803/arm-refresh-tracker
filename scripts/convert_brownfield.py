"""Convert brownfield_service_status.xlsx -> data/brownfield.json.

One-time-ish script. Run from the repo root whenever the source xlsx
in azure-sdk-for-js changes substantially.

Usage:
    python scripts/convert_brownfield.py path/to/brownfield_service_status.xlsx
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl


REPO_ROOT = Path(__file__).resolve().parents[1]
OUTPUT = REPO_ROOT / "data" / "brownfield.json"

EXPECTED_HEADERS = {
    "service": "Service",
    "armNamespace": "ARM Namespace",
    "specFolder": "Spec Folder",
    "sdkPackageName": "SDK Package Name",
}


def _cell(value):
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def convert(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    headers = {cell.value: idx for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))}
    for key, header_name in EXPECTED_HEADERS.items():
        if header_name not in headers:
            raise SystemExit(f"missing column in xlsx: {header_name}")

    out: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c is not None for c in row):
            break
        record = {
            key: _cell(row[headers[header_name]])
            for key, header_name in EXPECTED_HEADERS.items()
        }
        if not record["service"] and not record["armNamespace"]:
            continue
        out.append(record)
    return out


def main() -> None:
    if len(sys.argv) != 2:
        print(__doc__)
        sys.exit(2)
    rows = convert(Path(sys.argv[1]))
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(rows, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    have_pkg = sum(1 for r in rows if r["sdkPackageName"])
    print(f"Wrote {OUTPUT.relative_to(REPO_ROOT)} ({len(rows)} rows, {have_pkg} with sdkPackageName).")


if __name__ == "__main__":
    main()
